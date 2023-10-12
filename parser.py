import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.utils.cell import get_column_letter

from errors import WorksheetDoesNotExist


def get_merged_cell_value(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng) != 0 else cell.value


def find_class_cell(sheet, number_columns, number_rows, value):
    for i in range(number_columns):
        for k in range(number_rows):
            cell = sheet[get_column_letter(i + 1) + str(k + 1)]
            if cell.value and value in str(cell.value):
                return cell.column, cell.row


def find_count_of_subjects(sheet, class_cell):
    count = 0
    for i in range(1, 20):
        if sheet.cell(column=class_cell[0], row=class_cell[1] + i).value is None and count != 0:
            break
        count += 1
    return count


def find_serial_numbers(sheet, class_cell):
    count = 0
    serial_numbers = []
    for i in range(1, 20):
        if sheet.cell(column=2, row=class_cell[1] + i).value is None and count != 0:
            break

        count += 1
        serial_numbers.append(count)

    return tuple(serial_numbers)


    def find_time_cells(sheet, class_cell):
        time_cells = []
        for i in range(1, 20):
            time_cells.append(sheet.cell(column=3, row=class_cell[1] + i).value)
            if sheet.cell(column=3, row=class_cell[1] + i).value is None:
                break
        return tuple(time_cells)


def find_class_subjects(sheet, class_cell):
    subjects = []
    for i in range(1, find_count_of_subjects(sheet, class_cell) + 1):
        cell = sheet.cell(column=class_cell[0], row=class_cell[1] + i)
        if isinstance(cell, MergedCell):
            # print(get_merged_cell_value(sheet, cell))
            subjects.append(get_merged_cell_value(sheet, cell))
        else:
            # print(cell.value)
            subjects.append(cell.value)
    return tuple(subjects)


def create_a_schedule(sheet, class_cell):
    return zip(find_serial_numbers(sheet, class_cell), find_time_cells(sheet, class_cell),
               find_class_subjects(sheet, class_cell))


book = openpyxl.load_workbook('./data/timetable_september.xlsx')


def get_schedule(data, class_name):
    sheet = book[data]
    try:
        number_rows = sheet.max_row
        number_columns = sheet.max_column

        result = create_a_schedule(sheet, find_class_cell(sheet, number_rows, number_columns, class_name))
    except KeyError:
        raise WorksheetDoesNotExist

    return result


def main():
    data = input()
    class_name = input()

    # for answer in get_schedule(data, class_name):
    #     print(f'№{answer[0]} Time: {answer[1]} Subject: {answer[2]}')

    text = '\n'.join(
        f'№{answer[0]} Время: {answer[1]} Предмет: {answer[2]}' for answer in get_schedule(data, class_name))
    print(text)


if __name__ == '__main__':
    main()
